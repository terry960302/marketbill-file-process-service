#!/usr/bin/env python
# coding: utf-8
from datetime import datetime
from gspread import Spreadsheet, Client
import gspread
from pytz import timezone
import boto3
from botocore.client import BaseClient
from config import ACCESS_KEY_ID, ACCESS_SECRET_KEY, BUCKET_NAME
import os
from oauth2client.service_account import ServiceAccountCredentials
from gspread.cell import Cell
import requests
from math import ceil
from models import receipt_process_input as dto
from models.receipt_process_output import ReceiptProcessOutput
from typing import List
import time
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import shutil
from PyPDF2 import PdfWriter, PdfReader, PdfFileReader
import jpype
from pathlib import Path

# 엑셀을 사용하는 영수증 처리 서비스(Deprecated 이유)
## - 구글 스프레드 시트 : 10초 넘어가는 속도.
## - aspose.cells 라이브러리 : 100개 인스턴스화 라이센스 이슈로 프로덕션에 사용불가
## - excel -> html -> pdf : 양식 스타일링 지원이 잘안됨
@DeprecationWarning("ReceiptService를 사용바랍니다.")
class ReceiptServiceExcel:
    SCOPE = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    JSON_KEY_PATH = "../credential/eloquent-branch-369106-1e961c642259.json"
    GOOGLE_SPREAD_SHEET_ACCOUNT = 'floway.wholesale@gmail.com'
    # ROOT_DIR = os.path.abspath(os.curdir)
    ROOT_DIR = Path(__file__).parent.parent
    LOCAL_STORAGE_PATH = f'{ROOT_DIR if ROOT_DIR != "/" else ""}/tmp_storage/'
    EXPORT_FILE_FORMAT = ".pdf"
    EXCEL_FILE_FORMAT = ".xlsx"
    STORAGE_TYPE = "s3"
    REMOTE_STORAGE_DIR = "file-process-services-storage"
    REMOTE_STORAGE_URI_PREFIX = "https://marketbill-storage.s3.ap-northeast-2.amazonaws.com/"
    REMOTE_CDN_URI_PREFIX = "https://d23zpri05ibxyp.cloudfront.net"

    def __init__(self, data: dto.ReceiptProcessInput, receipt_form_file_name: str):
        print("root directory : ", ReceiptServiceExcel.ROOT_DIR)
        # 구글 계정 연동
        self.credential = None
        # s3에 저장될 파일명
        file_name: str = f"receipt_{data.wholesaler.name}_{data.orderNo}"
        self.file_name = file_name
        # pdf 파일 저장 경로
        self.local_file_path: str = ReceiptServiceExcel.LOCAL_STORAGE_PATH + file_name + ReceiptServiceExcel.EXPORT_FILE_FORMAT
        self.receipt_form_file_name = receipt_form_file_name
        self.order_no = data.orderNo
        self.retailer_name = data.retailer.name
        self.wholesaler_name = data.wholesaler.name
        self.order_items: List[dto.OrderItem] = self._filter_not_null(
            data.orderItems)
        return

    @staticmethod
    def _get_today():
        today = datetime.now(timezone('Asia/Seoul'))
        today = datetime.strftime(today, '%Y-%m-%d')
        return today

    @staticmethod
    def _filter_not_null(order_list: List[dto.OrderItem]) -> List[dto.OrderItem]:
        def is_not_null(item: dto.OrderItem):
            return item.price is not None

        return list(filter(is_not_null, order_list))

    @staticmethod
    def _get_profile() -> str:
        profile = os.environ.get("PROFILE")
        if profile is None:
            profile = "local"
        else:
            profile = str(profile)
        return profile

    ### ------------------------  local 엑셀 파일 처리 코드 ---------------------------------- ###

    # local
    def process_receipt_from_local(self) -> ReceiptProcessOutput:
        excel_file, wb = self.create_excel_from_receipt_data()
        print("1. EXCEL 파일 생성 완료")
        self.export_pdf_from_local_excel(excel_file)
        print("2. PDF 파일 변환 완료")
        metadata = self.upload_receipt_to_s3()
        print('3. S3 원격저장소 업로드 완료')
        return self.create_output(metadata)

    # local
    def create_excel_from_receipt_data(self) -> (str, Workbook):
        form_excel_path = ReceiptServiceExcel.LOCAL_STORAGE_PATH + self.receipt_form_file_name + ReceiptServiceExcel.EXCEL_FILE_FORMAT
        new_excel_path = ReceiptServiceExcel.LOCAL_STORAGE_PATH + self.file_name + ReceiptServiceExcel.EXCEL_FILE_FORMAT

        try:
            ## Copy formed sheet
            shutil.copy(form_excel_path, new_excel_path)
            wb: Workbook = load_workbook(new_excel_path)
            sheet: Worksheet = wb.active

            ## Update basic info
            sheet.cell(row=2, column=2, value=self.order_no)
            sheet.cell(row=2, column=7, value=self.retailer_name)

            ## Update tot_price
            cur_date = ReceiptServiceExcel._get_today()
            calc_prices = list(map(lambda item: item.price * item.quantity, self.order_items))
            tot_price = sum(calc_prices)
            sheet.cell(row=8, column=1, value=cur_date)
            # sheet.cell(row=8, column=4, value=tot_price)
            sheet.cell(row=8, column=5, value=tot_price)
            sheet.cell(row=24, column=11, value=tot_price)

            ## Update flower items
            self._update_flower_items(wb, sheet)

            # 수정사항 저장
            wb.save(new_excel_path)
        except Exception as e:
            if os.path.isfile(new_excel_path):
                os.remove(new_excel_path)
            raise

        return new_excel_path, wb

    # local
    def _update_flower_items(self, wb: Workbook, sheet: Worksheet):
        def update_rows(_sheet: Worksheet, row_idx: int, _item: dto.OrderItem):
            _sheet.cell(row=row_idx, column=1,
                        value=f'({_item.flower.flowerType.name}){_item.flower.name}-{_item.grade}')
            _sheet.cell(row=row_idx, column=5, value=_item.quantity)
            _sheet.cell(row=row_idx, column=7, value=_item.price)
            _sheet.cell(row=row_idx, column=11, value=_item.price * _item.quantity)

        ## Update items
        max_row = 13  # 한 영수증에 들어갈 수 있는 항목 수
        sheet_num = 1  # 워크시트 n번째
        start_row_idx = 11  # 항목이 들어가기 시작하는 행 n번째
        cur_sheet = sheet

        for i in range(0, len(self.order_items)):
            # 13번째 항목을 채우면 다음 시트로 이동
            if i != 0 and i % max_row == 0:
                sheet_num += 1
                sheet_name = f'Sheet{sheet_num}'
                copied_sheet = wb.copy_worksheet(cur_sheet)  # copy basic info section from sheet
                copied_sheet.title = sheet_name
                # 복제할 때 항목칸은 삭제
                for row in copied_sheet['A11:K23']:
                    for cell in row:
                        cell.value = None

                cur_sheet = copied_sheet

            item = self.order_items[i]

            # 첫번째 장만 존재하는 경우(굳이 시트 재할당 필요x)
            if sheet_num <= 1:
                row_idx = i + start_row_idx
                update_rows(cur_sheet, row_idx, item)
            else:
                # 시트가 넘어가면 다시 11행부터 cell에 넣어주기 위한 셋업
                row_idx = (i + start_row_idx) - (max_row * (sheet_num - 1))
                update_rows(cur_sheet, row_idx, item)
        return

    # local
    def export_pdf_from_local_excel(self, excel_file_path: str):
        import asposecells
        import asposecells.api as aspose

        size_to_reduce = 20
        pdf_file_path = self.local_file_path

        try:
            # load_options = aspose.LoadOptions(aspose.MemorySetting.MEMORY_PREFERENCE)
            workbook: aspose.Workbook = aspose.Workbook(excel_file_path)
            workbook.getSettings().setLocale(jpype.java.util.Locale.KOREA)
            save_options = aspose.PdfSaveOptions()
            save_options.setOnePagePerSheet(True)
            workbook.save(pdf_file_path, save_options)
            workbook.dispose()

            reader = PdfReader(pdf_file_path)
            writer = PdfWriter()

            for page in reader.pages:
                width, height = page.cropbox.upper_right
                page.cropbox.upperLeft = (0, height - size_to_reduce)
                page.cropbox.upperRight = (width, height - size_to_reduce)
                page.cropbox.lowerRight = (width, size_to_reduce)
                page.cropbox.lowerLeft = (0, size_to_reduce)
                writer.addPage(page)

            with open(pdf_file_path, 'wb') as fp:
                writer.write(fp)

        except Exception as e:
            raise
        finally:
            if os.path.isfile(excel_file_path):
                os.remove(excel_file_path)
        return

    # common
    def upload_receipt_to_s3(self) -> str:
        profile = self._get_profile()
        object_name = f'{ReceiptServiceExcel.REMOTE_STORAGE_DIR}/{profile}/{self.file_name}{ReceiptServiceExcel.EXPORT_FILE_FORMAT}'

        s3_client: BaseClient = boto3.client(
            ReceiptServiceExcel.STORAGE_TYPE,
            aws_access_key_id=ACCESS_KEY_ID,
            aws_secret_access_key=ACCESS_SECRET_KEY
        )

        response = s3_client.upload_file(
            self.local_file_path,
            BUCKET_NAME,
            object_name)

        pdf = PdfFileReader(open(self.local_file_path, "rb"))
        pdf_info = pdf.getDocumentInfo()
        pdf_metadata = str(pdf_info)

        if os.path.isfile(self.local_file_path):
            os.remove(self.local_file_path)

        return pdf_metadata

    # common
    def create_output(self, pdf_metadata: str) -> ReceiptProcessOutput:
        # [Warning] remote_dir 경로를 제외하고 요청을 보내야 CDN 을 통해 파일을 읽을 수 있습니다.(CDN 의 origin 소스 경로 설정으로 인해.)
        profile = self._get_profile()
        cached_file_path = f'{ReceiptServiceExcel.REMOTE_CDN_URI_PREFIX}/{profile}/{self.file_name}{ReceiptServiceExcel.EXPORT_FILE_FORMAT}'
        return ReceiptProcessOutput(
            file_name=self.file_name,
            file_format=ReceiptServiceExcel.EXPORT_FILE_FORMAT,
            file_path=cached_file_path,
            metadata=pdf_metadata
        )

    ### ------------------------  gspread 원격 처리 코드 ---------------------------------- ###

    # gspread
    def process_receipt_from_google_spreadsheet(self) -> ReceiptProcessOutput:
        doc = self.get_formed_gspreadsheet_doc()
        print('1. 기본 세팅 완료(영수증 양식 엑셀 준비)')
        self.update_data_to_gspreadsheet(doc)
        print('2. 스프레드시트 내 데이터 입력 완료')
        self.export_pdf_from_gspreadsheet(doc)
        print('3. Spreadsheet -> local pdf 추출 완료')
        metadata = self.upload_receipt_to_s3()
        print('4. local pdf -> s3 업로드 완료')
        return self.create_output(metadata)

    # gspread
    def get_formed_gspreadsheet_doc(self):
        # authenticate
        credential = ServiceAccountCredentials.from_json_keyfile_name(ReceiptServiceExcel.JSON_KEY_PATH,
                                                                      ReceiptServiceExcel.SCOPE)
        self.credential = credential
        gc: Client = gspread.authorize(credential)

        # create spreadsheet using pre-created-form
        doc: Spreadsheet = gc.open(self.receipt_form_file_name)
        new_spreadsheet = gc.copy(doc.id, title=self.file_name)
        new_spreadsheet.share(ReceiptServiceExcel.GOOGLE_SPREAD_SHEET_ACCOUNT, perm_type='user', role='writer')
        doc = gc.open(self.file_name)
        return doc

    # gspread
    def create_gspreadsheet_cells(self, d: dto.OrderItem, num: int, cells):
        flower_name = d.flower.name
        flower_type = d.flower.flowerType.name

        grade = d.grade
        qty = d.quantity
        prc = int(d.price)
        tot_prc = int(prc * qty)

        item = "(" + flower_type + ")" + flower_name + "-" + grade

        cell_item = Cell(num, 1, item)
        cell_qty = Cell(num, 5, qty)
        cell_prc = Cell(num, 7, prc)
        cell_tot_prc = Cell(num, 11, tot_prc)

        cell = [cell_item, cell_qty, cell_prc, cell_tot_prc]

        cells = cells + cell

        return cells

    # gspread
    def update_basic_info_to_gspreadsheet(self, sheet: gspread.Worksheet):
        # 기본 정보 및 판매 합계 입력
        order_id = Cell(2, 2, self.order_no)
        retail_name = Cell(2, 7, self.retailer_name)
        issue_date = Cell(8, 1, ReceiptServiceExcel._get_today())

        calc_prices = list(map(lambda item: item.price *
                                            item.quantity, self.order_items))
        tot_price = sum(calc_prices)

        top_tot_price = Cell(8, 4, tot_price)
        bottom_tot_price = Cell(24, 11, tot_price)

        sheet.update_cells(
            [order_id, retail_name, issue_date, top_tot_price, bottom_tot_price])
        return

    # gspread
    def export_pdf_from_gspreadsheet(self, doc):
        url = 'https://docs.google.com/spreadsheets/export?format=pdf&size=statement&gridlines=false&scale=3&horizontal_alignment=CENTER&vertical_alignment=CENTER&id=' + doc.id
        headers = {'Authorization': 'Bearer ' + self.credential.create_delegated("").get_access_token().access_token}
        res = requests.get(url, headers=headers)

        with open(self.local_file_path, 'wb') as f:
            f.write(res.content)

        print('finish exporting pdf')

    # gspread
    def update_data_to_gspreadsheet(self, doc: Spreadsheet):
        max_row = 13
        # max_row 개수보다 많으면 영수증을 분리함.
        receipt_num = ceil(len(self.order_items) / max_row)

        default_sheet_name = "Sheet1"
        sheet: gspread.Worksheet = doc.worksheet(default_sheet_name)

        for n in range(receipt_num):
            # 첫 영수증에는 기본 정보 입력(+ 주문항목 입력)
            if n == 0:

                # 기본 정보 입력
                self.update_basic_info_to_gspreadsheet(sheet)

                cells = []

                end = (n + 1) * max_row

                for i, d in enumerate(self.order_items[:end]):
                    num = i + 11
                    cells = self.create_gspreadsheet_cells(d, num, cells)  # 주문 항목 입력

                sheet.update_cells(cells)

            # 그 다음 장(paper)부터는 기본 정보가 안 들어감
            elif n > 0:

                name = f'Sheet{str(n + 1)}'

                new_sheet: gspread.Worksheet = sheet.duplicate(insert_sheet_index=n, new_sheet_name=name)
                new_sheet.batch_clear(['A11:K23'])

                cells = []

                start = n * max_row
                end = (n + 1) * max_row

                for i, d in enumerate(self.order_items[start:end]):
                    num = i + 11
                    cells = self.create_gspreadsheet_cells(d, num, cells)

                new_sheet.update_cells(cells)
        return

    def excel_to_html_pdf(self):
        # create excel
        excel_path, workbook = self.create_excel_from_receipt_data()

        # excel to html
        from xlsx2html import xlsx2html
        import io

        xlsx_file = open(excel_path, 'rb')
        out_file = io.StringIO()
        xlsx2html(xlsx_file, out_file, locale='ko')
        out_file.seek(0)
        html_content = out_file.read()

        # html to pdf
        import pdfkit

        pdf_path = ReceiptServiceExcel.LOCAL_STORAGE_PATH + self.file_name + ".pdf"
        options = {
            'zoom': 1.2,
            'minimum-font-size': 16,
            'page-size': 'A5',
            'margin-top': '0.38in',
            'margin-bottom': '0.38in',
            'margin-left': '0.7in',
            'margin-right': '0.7in',
            'encoding': "ko_KR.UTF-8",
            'custom-header': [
                ('Accept-Encoding', 'gzip')
            ],
            'cookie': [],
            'no-outline': None
        }
        pdfkit.from_string(html_content, pdf_path, options=options)
        return

# def _create_json_mock(num: int = 100) -> dict:
#     today = datetime.now(timezone('Asia/Seoul'))
#     today = datetime.strftime(today, '%Y%m%d')
#     basic_info = {
#         "orderNo": f'{today}M1230918',
#         "retailer": {
#             "name": "꽃소매"
#         },
#         "wholesaler": {
#             "name": "꽃도매"
#         },
#     }
#
#     items = []
#     for i in range(0, num):
#         item = {
#             "flower": {
#                 "name": "테데오옐로우",
#                 "flowerType": {
#                     "name": "국화"
#                 }
#             },
#             "quantity": 17,
#             "grade": "상",
#             "price": 10000
#         }
#         item["flower"]["name"] = f'랜덤꽃{i}'
#         item["price"] = random.randrange(1000, 10000)
#         item["quantity"] = random.randrange(10, 100)
#         items.append(item)
#
#     basic_info["orderItems"] = items
#     return basic_info

#
#
# if __name__ == "__main__":
#     def init_jvm():
#         if jpype.isJVMStarted():
#             return
#         stream = os.popen('java -version')
#         output = stream.read()
#         print('JVM path : ', jpype.getDefaultJVMPath())
#         jpype.startJVM(jpype.getDefaultJVMPath())
#         jpype.java.lang.System.out.println("JVM checked from java functions")
#         print(output)
#
#
#     init_jvm()
#     print("init")
#     start_time = time.time()
#     json_object = _create_json_mock(num=20)
#
#     receipt_form_name = 'receipt_001'
#
#     json_input = dto.ReceiptProcessInput(**json_object)
#     print("총 상품 개수(None포함) : ", len(json_input.orderItems))
#     services = ReceiptService(json_input, receipt_form_name)
#     # output = services.process_receipt_from_google_spreadsheet() # 구글 스프레드시트 버전
#     output = services.process_receipt_from_local()  # 로컬 엑셀파일 처리 버전
#
#     print("결과")
#     print(output.to_dict())
#     print()
#     print("총 소요시간 --- %s seconds ---" % (time.time() - start_time))
