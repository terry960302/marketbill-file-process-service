#!/usr/bin/env python
# coding: utf-8
from datetime import datetime
from gspread import Spreadsheet, Client
import gspread
from pytz import timezone
import boto3
from botocore.client import BaseClient
from config import ACCESS_KEY_ID, ACCESS_SECRET_KEY, BUCKET_NAME
import os
from oauth2client.service_account import ServiceAccountCredentials
from gspread.cell import Cell
import requests
from math import ceil
from model import receipt_process_input as dto
from model.receipt_process_output import ReceiptProcessOutput
from typing import List
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
import time
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import shutil
import jpype
import asposecells
jpype.startJVM()
import asposecells.api as aspose
from PyPDF2 import PdfWriter, PdfReader, PdfMerger


class ReceiptService:
    SCOPE = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    JSON_KEY_PATH = "../credential/eloquent-branch-369106-1e961c642259.json"
    GOOGLE_SPREAD_SHEET_ACCOUNT = 'floway.wholesale@gmail.com'
    LOCAL_STORAGE_PATH = '../tmp_storage/'
    EXPORT_FILE_FORMAT = ".pdf"
    EXCEL_FILE_FORMAT = ".xlsx"
    STORAGE_TYPE = "s3"
    REMOTE_STORAGE_DIR = "file-process-service-storage"
    REMOTE_STORAGE_URI_PREFIX = "https://marketbill-storage.s3.ap-northeast-2.amazonaws.com/"
    REMOTE_CDN_URI_PREFIX = "https://d23zpri05ibxyp.cloudfront.net"

    def __init__(self, data: dto.ReceiptProcessInput, receipt_form_file_name: str):
        # 구글 계정 연동
        self.credential = None
        # s3에 저장될 파일명
        file_name: str = f"receipt_{data.wholesaler.name}_" + str(data.orderNo)
        self.file_name = file_name
        # pdf 파일 저장 경로
        self.local_file_path: str = ReceiptService.LOCAL_STORAGE_PATH + file_name + ReceiptService.EXPORT_FILE_FORMAT
        self.receipt_form_file_name = receipt_form_file_name
        self.order_no = data.orderNo
        self.retailer_name = data.retailer.name
        self.wholesaler_name = data.wholesaler.name
        self.order_items: List[dto.OrderItem] = self._filter_not_null(
            data.orderItems)
        return

    @staticmethod
    def get_today():
        today = datetime.now(timezone('Asia/Seoul'))
        today = datetime.strftime(today, '%Y-%m-%d')
        return today

    @staticmethod
    def _filter_not_null(order_list: List[dto.OrderItem]) -> List[dto.OrderItem]:
        def is_not_null(item: dto.OrderItem):
            return item.price is not None

        return list(filter(is_not_null, order_list))

    @staticmethod
    def _get_profile() -> str:
        profile = os.environ.get("PROFILE")
        if profile is None:
            profile = "local"
        else:
            profile = str(profile)
        return profile

    ### ------------------------  local 엑셀 파일 처리 코드 ---------------------------------- ###

    # local
    def process_receipt_from_local(self) -> ReceiptProcessOutput:
        excel_file = self.create_excel_from_receipt_data()
        print("1. 엑셀 파일 제작 완료")
        self.export_pdf_from_local_excel(excel_file)
        print("2. pdf 파일 추출 완료")
        metadata = self.upload_receipt_to_s3()
        print('3. local pdf -> s3 업로드 완료')
        return self.create_output(metadata)

    # local
    def create_excel_from_receipt_data(self) -> str:
        form_excel_path = ReceiptService.LOCAL_STORAGE_PATH + self.receipt_form_file_name + ReceiptService.EXCEL_FILE_FORMAT
        new_excel_path = ReceiptService.LOCAL_STORAGE_PATH + self.file_name + ReceiptService.EXCEL_FILE_FORMAT

        try:
            shutil.copy(form_excel_path, new_excel_path)
            wb: Workbook = load_workbook(new_excel_path)
            sheet: Worksheet = wb.active

            ## Update basic info
            sheet.cell(row=2, column=2, value=self.order_no)
            sheet.cell(row=2, column=7, value=self.retailer_name)

            ## Update tot_price
            cur_date = ReceiptService.get_today()
            calc_prices = list(map(lambda item: item.price * item.quantity, self.order_items))
            tot_price = sum(calc_prices)
            sheet.cell(row=8, column=1, value=cur_date)
            sheet.cell(row=8, column=4, value=tot_price)
            sheet.cell(row=24, column=11, value=tot_price)

            def update_rows(_sheet: Worksheet, row_idx: int, _item: dto.OrderItem):
                _sheet.cell(row=row_idx, column=1,
                            value=f'({_item.flower.flowerType.name}){_item.flower.name}-{_item.grade}')
                _sheet.cell(row=row_idx, column=5, value=_item.quantity)
                _sheet.cell(row=row_idx, column=7, value=_item.price)
                _sheet.cell(row=row_idx, column=11, value=_item.price * _item.quantity)

            ## Update items
            max_row = 13  # 한 영수증에 들어갈 수 있는 항목 수
            sheet_num = 1  # 워크시트 n번째
            start_row_idx = 11  # 항목이 들어가기 시작하는 행 n번째
            cur_sheet = sheet

            for i in range(0, len(self.order_items)):
                # 13번째 항목을 채우면 다음 시트로 이동
                if i != 0 and i % max_row == 0:
                    sheet_num += 1
                    sheet_name = f'Sheet{sheet_num}'
                    copied_sheet = wb.copy_worksheet(cur_sheet)  # copy basic info section from sheet
                    copied_sheet.title = sheet_name
                    # 복제할 때 항목칸은 삭제
                    for row in copied_sheet['A11:K23']:
                        for cell in row:
                            cell.value = None

                    cur_sheet = copied_sheet

                item = self.order_items[i]

                # 첫번째 장만 존재하는 경우(굳이 시트 재할당 필요x)
                if sheet_num <= 1:
                    row_idx = i + start_row_idx
                    update_rows(cur_sheet, row_idx, item)
                else:
                    # 시트가 넘어가면 다시 11행부터 cell에 넣어주기 위한 셋업
                    row_idx = (i + start_row_idx) - (max_row * (sheet_num - 1))
                    update_rows(cur_sheet, row_idx, item)

            # 수정사항 저장
            wb.save(new_excel_path)
        except Exception as e:
            if os.path.isfile(new_excel_path):
                os.remove(new_excel_path)
            raise

        return new_excel_path

    # local
    def export_pdf_from_local_excel(self, excel_file_path: str):
        size_to_reduce = 20
        pdf_file_path = self.local_file_path

        try:
            workbook = aspose.Workbook(excel_file_path)
            save_options = aspose.PdfSaveOptions()
            save_options.setOnePagePerSheet(True)
            workbook.save(pdf_file_path, save_options)
            jpype.shutdownJVM()

            reader = PdfReader(pdf_file_path)
            writer = PdfWriter()

            for page in reader.pages:
                width, height = page.cropbox.upper_right
                page.cropbox.upperLeft = (0, height - size_to_reduce)
                page.cropbox.upperRight = (width, height - size_to_reduce)
                page.cropbox.lowerRight = (width, size_to_reduce)
                page.cropbox.lowerLeft = (0, size_to_reduce)
                writer.addPage(page)

            with open(pdf_file_path, 'wb') as fp:
                writer.write(fp)

            print("export_as_pdf completed!")
        except Exception as e:
            raise
        finally:
            if os.path.isfile(excel_file_path):
                os.remove(excel_file_path)
        return

    # common
    def upload_receipt_to_s3(self) -> str:  # f = 파일명
        profile = self._get_profile()
        object_name = f'{ReceiptService.REMOTE_STORAGE_DIR}/{profile}/{self.file_name}{ReceiptService.EXPORT_FILE_FORMAT}'

        s3_client: BaseClient = boto3.client(
            ReceiptService.STORAGE_TYPE,
            aws_access_key_id=ACCESS_KEY_ID,
            aws_secret_access_key=ACCESS_SECRET_KEY
        )

        response = s3_client.upload_file(
            self.local_file_path,
            BUCKET_NAME,
            object_name)

        fp = open(self.local_file_path, 'rb')
        pdf_parser = PDFParser(fp)
        pdf_doc = PDFDocument(pdf_parser)
        pdf_metadata = str(pdf_doc.info)

        if os.path.isfile(self.local_file_path):
            os.remove(self.local_file_path)
        return pdf_metadata

    # common
    def create_output(self, pdf_metadata: str) -> ReceiptProcessOutput:
        # [Warning] remote_dir 경로를 제외하고 요청을 보내야 CDN 을 통해 파일을 읽을 수 있습니다.(CDN 의 origin 경로 설정으로 인해.)
        profile = self._get_profile()
        cached_file_path = f'{ReceiptService.REMOTE_CDN_URI_PREFIX}/{profile}/{self.file_name}{ReceiptService.EXPORT_FILE_FORMAT}'
        return ReceiptProcessOutput(
            file_name=self.file_name,
            file_format=ReceiptService.EXPORT_FILE_FORMAT,
            file_path=cached_file_path,
            metadata=pdf_metadata
        )

    ### ------------------------  gspread 원격 처리 코드 ---------------------------------- ###

    # gspread
    def process_receipt_from_google_spreadsheet(self) -> ReceiptProcessOutput:
        doc = self.get_formed_gspreadsheet_doc()
        print('기본 세팅 완료(영수증 양식 엑셀 준비)')
        self.update_data_to_gspreadsheet(doc)
        print('스프레드시트 내 데이터 입력 완료')
        self.export_pdf_from_gspreadsheet(doc)
        print('Spreadsheet -> local pdf 추출 완료')
        metadata = self.upload_receipt_to_s3()
        print('local pdf -> s3 업로드 완료')
        return self.create_output(metadata)

    # gspread
    def get_formed_gspreadsheet_doc(self):
        # authenticate
        credential = ServiceAccountCredentials.from_json_keyfile_name(ReceiptService.JSON_KEY_PATH,
                                                                      ReceiptService.SCOPE)
        self.credential = credential
        gc: Client = gspread.authorize(credential)

        # create spreadsheet using pre-created-form
        doc: Spreadsheet = gc.open(self.receipt_form_file_name)
        new_spreadsheet = gc.copy(doc.id, title=self.file_name)
        new_spreadsheet.share(ReceiptService.GOOGLE_SPREAD_SHEET_ACCOUNT, perm_type='user', role='writer')
        doc = gc.open(self.file_name)
        return doc

    # gspread
    def create_gspreadsheet_cells(self, d: dto.OrderItem, num: int, cells):
        flower_name = d.flower.name
        flower_type = d.flower.flowerType.name

        grade = d.grade
        qty = d.quantity
        prc = int(d.price)
        tot_prc = int(prc * qty)

        item = "(" + flower_type + ")" + flower_name + "-" + grade

        cell_item = Cell(num, 1, item)
        cell_qty = Cell(num, 5, qty)
        cell_prc = Cell(num, 7, prc)
        cell_tot_prc = Cell(num, 11, tot_prc)

        cell = [cell_item, cell_qty, cell_prc, cell_tot_prc]

        cells = cells + cell

        return cells

    # gspread
    def update_basic_info_to_gspreadsheet(self, sheet: gspread.Worksheet):
        # 기본 정보 및 판매 합계 입력
        order_id = Cell(2, 2, self.order_no)
        retail_name = Cell(2, 7, self.retailer_name)
        issue_date = Cell(8, 1, ReceiptService.get_today())

        calc_prices = list(map(lambda item: item.price *
                                            item.quantity, self.order_items))
        tot_price = sum(calc_prices)

        top_tot_price = Cell(8, 4, tot_price)
        bottom_tot_price = Cell(24, 11, tot_price)

        sheet.update_cells(
            [order_id, retail_name, issue_date, top_tot_price, bottom_tot_price])
        return

    # gspread
    def export_pdf_from_gspreadsheet(self, doc):
        url = 'https://docs.google.com/spreadsheets/export?format=pdf&size=statement&gridlines=false&scale=3&horizontal_alignment=CENTER&vertical_alignment=CENTER&id=' + doc.id
        headers = {'Authorization': 'Bearer ' + self.credential.create_delegated("").get_access_token().access_token}
        res = requests.get(url, headers=headers)

        with open(self.local_file_path, 'wb') as f:
            f.write(res.content)

        print('finish exporting pdf')

    # gspread
    def update_data_to_gspreadsheet(self, doc: Spreadsheet):
        max_row = 13
        # max_row 개수보다 많으면 영수증을 분리함.
        receipt_num = ceil(len(self.order_items) / max_row)

        default_sheet_name = "Sheet1"
        sheet: gspread.Worksheet = doc.worksheet(default_sheet_name)

        for n in range(receipt_num):
            # 첫 영수증에는 기본 정보 입력(+ 주문항목 입력)
            if n == 0:

                # 기본 정보 입력
                self.update_basic_info_to_gspreadsheet(sheet)

                cells = []

                end = (n + 1) * max_row

                for i, d in enumerate(self.order_items[:end]):
                    num = i + 11
                    cells = self.create_gspreadsheet_cells(d, num, cells)  # 주문 항목 입력

                sheet.update_cells(cells)

            # 그 다음 장(paper)부터는 기본 정보가 안 들어감
            elif n > 0:

                name = f'Sheet{str(n + 1)}'

                new_sheet: gspread.Worksheet = sheet.duplicate(insert_sheet_index=n, new_sheet_name=name)
                new_sheet.batch_clear(['A11:K23'])

                cells = []

                start = n * max_row
                end = (n + 1) * max_row

                for i, d in enumerate(self.order_items[start:end]):
                    num = i + 11
                    cells = self.create_gspreadsheet_cells(d, num, cells)

                new_sheet.update_cells(cells)
        return

    # local
    # 읽기 편한 openpyxl 데이터 입력 코드입니다.(대신 효율성이 별로...물론 데이터가 많은 경우에 해당)
    @DeprecationWarning
    def _create_new_excel(self, wb: Workbook, sheet: Worksheet):
        max_row = 13  # 한 영수증에 들어갈 수 있는 항목 수
        # 가독용 코드
        sheet_count = ceil(len(self.order_items) / max_row)
        for i in range(0, sheet_count):
            if i == 0:
                self._update_items_on_cells(sheet, self.order_items[i * max_row: (i + 1) * max_row])
            else:
                sheet_num = i + 1
                sheet_name = f'Sheet{sheet_num}'
                copied_sheet = wb.copy_worksheet(sheet)  # copy basic info section from sheet
                copied_sheet.title = sheet_name
                self._update_items_on_cells(copied_sheet, self.order_items[i * max_row: (i + 1) * max_row])

    # local
    # 읽기 편한 openpyxl 데이터 입력 코드입니다.(대신 효율성이 별로...물론 데이터가 많은 경우에 해당)
    @DeprecationWarning
    def _update_items_on_cells(self, sheet: Worksheet, items: List[dto.OrderItem]):
        min_row_idx = 11
        for i in range(0, len(items)):
            item = items[i]
            row_idx = i + min_row_idx
            sheet.cell(row=row_idx, column=1, value=f'({item.flower.flowerType.name}){item.flower.name}-{item.grade}')
            sheet.cell(row=row_idx, column=5, value=item.quantity)
            sheet.cell(row=row_idx, column=7, value=item.price)
            sheet.cell(row=row_idx, column=11, value=item.price * item.quantity)


# test
if __name__ == "__main__":
    print("init")
    start_time = time.time()
    json_object = {
        "orderNo": "20221209M1230918",
        "retailer": {
            "name": "꽃소매"
        },
        "wholesaler": {
            "name": "꽃도매"
        },
        "orderItems": [
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우3",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 390,
                "grade": "상",
                "price": 3404
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            }, {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            }, {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },
            {
                "flower": {
                    "name": "테데오옐로우",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 17,
                "grade": "상",
                "price": 10000
            },

            {
                "flower": {
                    "name": "테데오옐로우2",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 123,
                "grade": "상",
                "price": 39450
            },
            {
                "flower": {
                    "name": "신명",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 23,
                "grade": "상",
                "price": None
            },
            {
                "flower": {
                    "name": "상그릴라",
                    "flowerType": {
                        "name": "국화"
                    }
                },
                "quantity": 84,
                "grade": "상",
                "price": None
            }
        ]
    }

    receipt_form_name = 'receipt_001'

    json_input = dto.ReceiptProcessInput(**json_object)
    print("총 상품 개수(None포함) : ", len(json_input.orderItems))
    service = ReceiptService(json_input, receipt_form_name)
    # output = service.process_receipt_from_google_spreadsheet() # 구글 스프레드시트 버전
    output = service.process_receipt_from_local()  # 로컬 엑셀파일 처리 버전

    print("결과")
    print(output.to_dict())
    print()
    print("총 소요시간 --- %s seconds ---" % (time.time() - start_time))
